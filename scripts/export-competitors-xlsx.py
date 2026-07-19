#!/usr/bin/env python3
"""export-competitors-xlsx.py
输出竞对网站分析 Excel：竞对 / 类型 / 月访问量(Semrush) / 全球排名 / 中英文标记。
流量数据来源：semrush.com/website/<domain>/overview/ 公开页（抓取日期 2026-07-19）。
输出：gsc/0719/competitor-traffic-0719.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "gsc" / "0719" / "competitor-traffic-0719.xlsx"

# (域名, 类型, 月访问量(万, None=无数据), 原始值, 数据月份, 全球排名, 中英文, 备注)
ROWS = [
    # 直接竞对：AI 占卜/命理工具站
    ("tarotap.com", "AI占卜工具站（直接竞对）", 78.9, "789.21K", "2026-06", 56360, "多语言（英为主）", "AI塔罗+多AI师；与本站 tarot 模块对位"),
    ("tarotoo.com", "AI占卜工具站（直接竞对）", 39.8, "398.04K", "2026-06", 98861, "多语言（英为主）", "免费AI塔罗+通灵问答"),
    ("tianjiyao.com", "AI占卜工具站（直接竞对）", None, "无数据", "", None, "中文为主（多语言）", "易经/八字/紫微/塔罗；低于 Semrush 阈值"),
    ("lingyuan.ai", "AI命理工具站（直接竞对）", None, "无数据", "", None, "中文", "八字/紫微/姓名；低于 Semrush 阈值"),
    ("suanmingzhun.com", "AI命理工具站（直接竞对）", None, "无数据", "", None, "中文", "八字/紫微/易经卜卦"),
    ("labyrinthos.co", "塔罗学习App", 269.0, "2.69M", "2026-06", 19293, "英文", "塔罗学习+占卜，App 为主"),
    ("tarot.com", "综合占卜内容站", 218.0, "2.18M", "2026-06", 23334, "英文", "老牌塔罗/占星内容站"),
    ("free-tarot-reading.net", "免费塔罗工具站", 68.4, "684.45K", "2026-06", None, "英文", "Semrush 显示为 tarotoo 近邻竞对"),
    ("evatarot.net", "免费塔罗工具站", 36.5, "364.73K", "2025-07", None, "英文/法文", "数据较旧，仅供参考"),
    # 西方占星头部
    ("astro-seek.com", "占星排盘工具站", 1217.0, "12.17M", "2026-06", 4261, "多语言（英为主）", "排盘工具流量天花板"),
    ("horoscope.com", "星座内容站", 1062.0, "10.62M", "2026-06", 4956, "英文", ""),
    ("astro.com", "占星排盘工具站", 942.0, "9.42M", "2026-06", 5630, "多语言", "Astrodienst，行业老牌"),
    ("cafeastrology.com", "占星内容站", 606.0, "6.06M", "2026-06", None, "英文", ""),
    ("astrology.com", "综合占星内容站", 339.0, "3.39M", "2026-06", 15622, "英文", ""),
    ("asknebula.com", "占星+真人咨询App", 95.7, "956.65K", "2026-06", 47822, "英文", "Nebula，投放榜单软文较多"),
    ("costarastrology.com", "AI占星App", 67.0, "669.5K", "2026-06", None, "英文", "Co-Star，App 为主，站点流量仅部分"),
    ("chani.com", "占星App+内容", 23.3, "232.81K", "2025-09", 152262, "英文", "数据较旧，仅供参考"),
    ("thepattern.com", "AI占星App", None, "无数据", "", None, "英文", "App 为主，站点低于阈值"),
    ("sanctuaryastrology.com", "占星App", None, "无数据", "", None, "英文", ""),
    # 真人占卜 marketplace
    ("astroyogi.com", "真人占星咨询marketplace", 184.0, "1.84M", "2026-06", 27170, "英文/印地语", "印度系"),
    ("astrotalk.com", "真人占星咨询marketplace", 168.0, "1.68M", "2026-06", 29423, "英文/印地语", "印度系"),
    ("keen.com", "真人占卜marketplace", 85.9, "859.24K", "2026-02", 48451, "英文", "数据较旧，仅供参考"),
    ("californiapsychics.com", "真人占卜marketplace", 44.6, "446.4K", "2026-06", None, "英文", ""),
    ("kasamba.com", "真人占卜marketplace", 11.0, "110.37K", "2025-11", None, "英文", "数据较旧，仅供参考"),
    ("purplegarden.co", "真人占卜marketplace", 9.0, "90.2K", "2026-06", 318637, "英文", ""),
    # 中文传统命理内容站
    ("k366.com", "中文命理内容站（华易网）", None, "无数据", "", None, "中文", "Semrush 对中文站/百度流量覆盖差，无数据≠无流量"),
    ("d1xz.net", "中文星座内容站（第一星座网）", None, "无数据", "", None, "中文", "同上"),
    ("xzw.com", "中文星座内容站（星座屋）", None, "无数据", "", None, "中文", "同上"),
    # 本站基准
    ("aiastrum.com", "本站（基准）", None, "无数据", "", None, "多语言", "新站，低于 Semrush 阈值"),
]

HEADER_FILL = PatternFill("solid", fgColor="4F39FA")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SELF_FILL = PatternFill("solid", fgColor="FFF3CD")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "竞对流量对比"
    ws.append(["竞对网站", "类型", "月访问量(万)", "Semrush原始值", "数据月份", "全球排名", "中英文站", "备注"])
    for row in sorted(ROWS, key=lambda r: -(r[2] or 0)):
        ws.append(list(row))
    for i, w in enumerate([22, 30, 12, 14, 10, 10, 16, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "aiastrum.com":
            for c in range(1, 9):
                ws.cell(r, c).fill = SELF_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    note = wb.create_sheet("说明")
    notes = [
        "数据来源：Semrush 公开流量页 semrush.com/website/<domain>/overview/，抓取于 2026-07-19。",
        "月访问量为 Semrush 估算的全部设备访问量，非精确值；不同月份数据并存已单独标注。",
        "「无数据」= 该站流量低于 Semrush 公开页展示阈值。中文站因 Semrush 对百度流量覆盖差，缺失不代表没有流量。",
        "App 类产品（Co-Star/The Pattern/Sanctuary/测测等）主要流量在应用内，站点访问量不能反映其真实规模。",
        "Keywords Everywhere API（2 credits/域名）可补充精确 organic traffic 估算，但当前账户无 credits。",
    ]
    for i, t in enumerate(notes, 1):
        note.cell(i, 1, t)
    note.column_dimensions["A"].width = 120

    wb.save(OUT)
    print(f"saved: {OUT}, {ws.max_row - 1} rows")


if __name__ == "__main__":
    main()
