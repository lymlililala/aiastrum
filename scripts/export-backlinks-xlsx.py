#!/usr/bin/env python3
"""export-backlinks-xlsx.py
把竞对外链渠道清单 + Serper 挖掘结果导出为 Excel。
输入：
  gsc/0719/backlink-prospects-0719.md          （人工整理的渠道清单）
  gsc/0719/competitor-backlinks-serper-0719.json（Serper 实跑原始数据）
输出：
  gsc/0719/backlink-prospects-0719.xlsx
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
GSC = ROOT / "gsc" / "0719"
MD = GSC / "backlink-prospects-0719.md"
JSON = GSC / "competitor-backlinks-serper-0719.json"
OPR = GSC / "openpagerank-0719.json"
OUT = GSC / "backlink-prospects-0719.xlsx"

CN_SLD = {"com", "edu", "gov", "net", "org", "ac"}


def registered_domain(url):
    host = re.sub(r"^https?://", "", url.strip()).split("/")[0].split(":")[0].lower()
    host = host.removeprefix("www.")
    parts = host.split(".")
    if len(parts) < 2 or not parts[-1]:
        return None
    if parts[-1] == "cn" and len(parts) >= 3 and parts[-2] in CN_SLD:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


def load_opr():
    if not OPR.exists():
        return {}
    return json.loads(OPR.read_text(encoding="utf8"))


def opr_fields(opr, url):
    """返回 (权重分, 全球排名, 引用域数) 或空。"""
    d = registered_domain(url) if url else None
    rec = opr.get(d) if d else None
    if not rec or not rec.get("found"):
        return (None, None, None)
    return (rec.get("opr"), rec.get("rank"), rec.get("ref_domains"))


def sort_key(opr_val):
    return -(opr_val if opr_val is not None else -1)

HEADER_FILL = PatternFill("solid", fgColor="4F39FA")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def parse_prospects_md():
    """从 markdown 清单提取 (分类, 名称, URL, 备注)。无链接的纯文本行记为无 URL 渠道。"""
    rows = []
    section = ""
    link_re = re.compile(r"\[([^\]]+)\]\((https?://[^)\s]+)\)")
    for raw in MD.read_text(encoding="utf8").splitlines():
        line = raw.strip()
        m = re.match(r"^(#{2,3})\s+(.*)", line)
        if m:
            section = re.sub(r"（.*?）", "", m.group(2)).strip()
            continue
        if not line.startswith("- "):
            continue
        body = line[2:]
        links = link_re.findall(body)
        note = link_re.sub("", body)
        note = re.sub(r"\s*/\s*", " / ", note).strip(" -—/")[:200]
        if links:
            for name, url in links:
                rows.append((section, name.strip(), url, note))
        else:
            # 无 markdown 链接的行（如 Reddit 版名、Profile 站列表），整行作为一条
            text = body.strip()
            if text and not text.startswith((">", "共 ")):
                rows.append((section, text[:80], "", ""))
    return rows


def main():
    wb = Workbook()
    opr = load_opr()

    # Sheet1: 渠道清单（按 OPR 权重降序）
    ws1 = wb.active
    ws1.title = "渠道清单"
    ws1.append(["分类", "名称", "URL", "OPR权重", "全球排名", "引用域数", "备注"])
    rows = []
    for section, name, url, note in parse_prospects_md():
        score, rank, refs = opr_fields(opr, url)
        rows.append((section, name, url, score, rank, refs, note))
    rows.sort(key=lambda r: sort_key(r[3]))
    for row in rows:
        ws1.append(list(row))
    style_sheet(ws1, [22, 40, 60, 10, 12, 10, 60])

    data = json.loads(JSON.read_text(encoding="utf8"))

    # Sheet2: 引用域名汇总（按 OPR 权重降序）
    ws2 = wb.create_sheet("引用域名汇总")
    ws2.append(["引用域名", "OPR权重", "全球排名", "链接的竞对", "页面类型", "页面数", "示例页"])
    ref_rows = []
    for r in data["refDomains"]:
        rec = opr.get(r["refDomain"]) or {}
        score = rec.get("opr") if rec.get("found") else None
        rank = rec.get("rank") if rec.get("found") else None
        ref_rows.append((r["refDomain"], score, rank, ", ".join(r["linksTo"]), "/".join(r["types"]), r["urlCount"], r["sampleUrl"]))
    ref_rows.sort(key=lambda r: sort_key(r[1]))
    for row in ref_rows:
        ws2.append(list(row))
    style_sheet(ws2, [30, 10, 12, 45, 20, 8, 70])

    # Sheet3: Serper 提及页明细（按提及页域名权重降序）
    ws3 = wb.create_sheet("提及页明细")
    ws3.append(["竞对域名", "提及页 URL", "OPR权重", "全球排名", "标题", "类型", "摘要"])
    page_rows = []
    for comp, pages in data["pages"].items():
        for p in pages:
            score, rank, _ = opr_fields(opr, p["url"])
            page_rows.append((comp, p["url"], score, rank, p["title"][:120], p["type"], p["snippet"][:200]))
    page_rows.sort(key=lambda r: sort_key(r[2]))
    for row in page_rows:
        ws3.append(list(row))
    style_sheet(ws3, [20, 70, 10, 12, 50, 12, 60])

    wb.save(OUT)
    print(f"saved: {OUT}")
    print(f"渠道清单 {ws1.max_row - 1} 行, 引用域名 {ws2.max_row - 1} 行, 提及页 {ws3.max_row - 1} 行")


if __name__ == "__main__":
    main()
